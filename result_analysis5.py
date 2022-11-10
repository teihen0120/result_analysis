#モジュールのインポート
import openpyxl as px
from openpyxl.styles import PatternFill
import os
import glob

#エクセルファイル読み込み
###############result_anarysysのフォルダのパスに書き換えてください##################
xlfile_folder_path = r"D:\D_tokudome\D_desktop\result_analysis\*.xlsx"
xlfile_path_list = glob.glob(xlfile_folder_path)

#色指定
fill_red = PatternFill(patternType="solid", fgColor="FF0000", bgColor="FF0000")
fill_green = PatternFill(patternType="solid", fgColor="008000", bgColor="008000")
fill_blue = PatternFill(patternType="solid", fgColor="0000FF", bgColor="0000FF")
#1だったらセルに日にちを入れる
for n, f in enumerate(xlfile_path_list):
    xlfile_path = f
    wb = px.load_workbook(xlfile_path)
    date_name = os.path.splitext(os.path.basename(xlfile_path))[0]
    sheet = wb["Sheet1"]
    #計測初日のファイル
    if n == 0:
        wb_init = wb
        sheet_init = wb_init["Sheet1"]
        print(date_name)
        for row in range(3, 301):
            for column in range(2, 22):
                if sheet_init.cell(row=row, column=column).value == None:
                    sheet_init.cell(row=row, column=column).value = 0
                elif sheet_init.cell(row=row, column=column).value == 1:
                    sheet.cell(row=row, column=column).fill = fill_red
                    sheet_init.cell(row=row, column=column).value = int(date_name)
    #計測２日目以降のファイル
    else:
        print(f)
        # sheet = wb["Sheet1"]
        for row in range(3, 301):
            for column in range(2, 22):
                value = sheet_init.cell(row=row, column=column).value
                if value == 0 and sheet.cell(row=row, column=column).value == 1:
                    sheet_init.cell(row=row, column=column).value = int(date_name)
                if sheet.cell(row=row, column=column).value == 1:
                    sheet.cell(row=row, column=column).fill = fill_blue

    wb.save(r"D:\D_tokudome\D_desktop\result\\" + str(date_name) + ".xlsx")
#出蕾日と開花日でブックを分割                    
wb_tsubomi = px.Workbook()
sheet_tsubomi = wb_tsubomi["Sheet"]

wb_flower = px.Workbook()
sheet_flower = wb_flower["Sheet"]

for row in range(1, 3):
    for column in range(1, 22):
        sheet_tsubomi.cell(row=row, column=column).value = sheet_init.cell(row=row, column=column).value
        sheet_flower.cell(row=row, column=column).value = sheet_init.cell(row=row, column=column).value
for row in range(3, 301):
    for column in range(1, 22):
        if row % 2 == 0:
            sheet_flower.cell(row=(row/2)+2, column=column).value = sheet_init.cell(row=row, column=column).value
        else:
            sheet_tsubomi.cell(row=(row//2)+2, column=column).value = sheet_init.cell(row=row, column=column).value

#保存







wb_init.save(r"D:\D_tokudome\D_desktop\result\result_" + str(date_name) + ".xlsx")
wb_tsubomi.save(r"D:\D_tokudome\D_desktop\result\result_tsubomi_" + str(date_name) + ".xlsx")
wb_flower.save(r"D:\D_tokudome\D_desktop\result\result_flower_" + str(date_name) + ".xlsx")

