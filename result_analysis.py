#モジュールのインポート
import openpyxl as px
import os

#エクセルファイル読み込み
xlfile_path = r"D:\D_tokudome\D_desktop\result_anarysys\20221017.xlsx"
wb = px.load_workbook(xlfile_path)

#ファイル名から日にち取得
date_name = os.path.splitext(os.path.basename(xlfile_path))[0]
print(date_name)

#エクセルファイルコピー
wb_copy = wb

#出蕾日行列取得
sheet = wb["Sheet1"]
print(sheet.cell(row=1, column=39).value)
sheet_copy = wb_copy["Sheet1"]
for row in range(3, 301, 2):
    for column in range(2, 22):
        value = sheet.cell(row=row, column=column).value
        if value == 1:
            sheet_copy.cell(row=row, column=column).value = int(date_name)
            print("row:", row, "column:", column, "\n")



#１を出蕾日に書き換え


#保存
# wb_copy.save("result.xlsx")
