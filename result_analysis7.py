#モジュールのインポート
import openpyxl as px
from openpyxl.styles import PatternFill
import os
import glob

#エクセルファイル読み込み
###############result_anarysysのフォルダのパスに書き換えてください##################
xlfile_folder_path = r"D:\D_tokudome\D_desktop\result_analysis\*.xlsx"
xlfile_path_list = glob.glob(xlfile_folder_path)

os.makedirs(r"D:\D_tokudome\D_desktop\result", exist_ok=True)
#色指定
fill_red = PatternFill(patternType="solid", fgColor="FF0000", bgColor="FF0000")
fill_green = PatternFill(patternType="solid", fgColor="008000", bgColor="008000")
fill_blue = PatternFill(patternType="solid", fgColor="0000FF", bgColor="0000FF")

wb_init = px.Workbook()
wb_init["Sheet"].title = "Sheet1"
sheet_init = wb_init["Sheet1"]

#1だったらセルに日にちを入れる
for n, f in enumerate(xlfile_path_list):
    print(f)
    wb = px.load_workbook(f)
    date_name = os.path.splitext(os.path.basename(f))[0]
    sheet = wb["Sheet1"]
    if n == 0:
        for row in range(1, 301):
            for column in range(1, 22):
                sheet_init.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
                if (3 <= row <= 300) and (2 <= column <= 21):
                    if sheet.cell(row=row, column=column).value == None:
                        sheet_init.cell(row=row, column=column).value = 0
                    elif sheet.cell(row=row, column=column).value == 1:
                        sheet_init.cell(row=row, column=column).value = int(date_name)
                        if row % 2 == 0:
                            sheet.cell(row=row, column=column).fill = fill_red
                        else:
                            sheet.cell(row=row, column=column).fill = fill_green
    else:
        for row in range(3, 301):
            for column in range(2, 22):
                value_init = sheet_init.cell(row=row, column=column).value
                value = sheet.cell(row=row, column=column).value
                if value == 1:
                    if row % 2 == 0:
                        sheet.cell(row=row, column=column).fill = fill_red
                    else:
                        sheet.cell(row=row, column=column).fill = fill_green
                    if value_init == 0:
                        sheet_init.cell(row=row, column=column).value = int(date_name)
                elif row % 2 == 0:
                    if not ((value_init and sheet_init.cell(row=row+1, column=column).value) == 0)\
                        and (sheet.cell(row=row+1, column=column).value == 0):
                            sheet.cell(row=row, column=column).fill = fill_blue
                # if ((not sheet_init.cell(row=(row/2)+2, column=column).value == 0 or not sheet_init.cell(row=(row//2)+2, column=column).value == 0)
                #     and (sheet.cell(row=(row/2)+2, column=column).value == 0 and sheet.cell(row=(row//2)+2, column=column).value == 0)):
                #     sheet.cell(row=row, column=column).fill = fill_blue



    wb.save(r"D:\D_tokudome\D_desktop\result\\" + str(date_name) + ".xlsx")

