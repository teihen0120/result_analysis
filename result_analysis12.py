#モジュールのインポート
import openpyxl as px
from openpyxl.styles import PatternFill
import os
import glob
import datetime

#エクセルファイル読み込み
###############エクセルが入っているフォルダのパスに書き換えてください##################
xlfile_folder_path = r"D:\D_tokudome\D_desktop\result_analysis\*.xlsx"
##################################################################################

#ファイル取得
xlfile_path_list = glob.glob(xlfile_folder_path)

###############保存先のパスに書き換えてください(tyousa, resultのところが保存ファイル名になっています)###################
#調査表保存用のパス
save_tyousa_path = r"D:\D_tokudome\D_desktop\tyousa"
#結果保存用のパス
save_result_path = r"D:\D_tokudome\D_desktop\result"
##################################################################################################################

#フォルダ生成
os.makedirs(save_tyousa_path, exist_ok=True)
os.makedirs(save_result_path, exist_ok=True)

#色指定
def get_red(plus_red=0):
    hex_red = f"{100+plus_red:02X}{0:02X}{0:02X}"
    red = PatternFill(patternType="solid", fgColor=hex_red, bgColor=hex_red)
    return red
def get_green(plus_green=0):
    hex_green = f"{0:02X}{100+plus_green:02X}{0:02X}"
    green = PatternFill(patternType="solid", fgColor=hex_green, bgColor=hex_green)
    return green
def get_blue(plus_blue=0):
    hex_blue = f"{0:02X}{0:02X}{100+plus_blue:02X}"
    blue = PatternFill(patternType="solid", fgColor=hex_blue, bgColor=hex_blue)
    return blue

#空のワークブック作成　result書き込み用
wb_init = px.Workbook()
wb_init["Sheet"].title = "Sheet1"
sheet_init = wb_init["Sheet1"]

wb_uekae_write = px.Workbook()
wb_uekae_write["Sheet"].title = "Sheet1"
sheet_uekae_write = wb_uekae_write["Sheet1"]

wb_pre = px.Workbook()
wb_pre["Sheet"].title = "Sheet1"
sheet_pre = wb_pre["Sheet1"]

wb_init_2 = px.Workbook()
wb_init_2["Sheet"].title = "Sheet1"
sheet_init_2 = wb_init_2["Sheet1"]

#出蕾日と開花日でブックを分割                    
wb_tsubomi = px.Workbook()
wb_tsubomi["Sheet"].title = "Sheet1"
sheet_tsubomi = wb_tsubomi["Sheet1"]

wb_flower = px.Workbook()
wb_flower["Sheet"].title = "Sheet1"
sheet_flower = wb_flower["Sheet1"]

wb_uekae = px.Workbook()
wb_uekae["Sheet"].title = "Sheet1"
sheet_uekae = wb_uekae["Sheet1"]

wb_tsubomi_2 = px.Workbook()
wb_tsubomi_2["Sheet"].title = "Sheet1"
sheet_tsubomi_2 = wb_tsubomi_2["Sheet1"]

wb_flower_2 = px.Workbook()
wb_flower_2["Sheet"].title = "Sheet1"
sheet_flower_2 = wb_flower_2["Sheet1"]

#出蕾日、開花日記録、色塗り
for n in range(0, len(xlfile_path_list)):
    print(xlfile_path_list[n])
    #現在と１つ前のワークブック作成
    wb = px.load_workbook(xlfile_path_list[n])
    sheet = wb["Sheet1"]
    
    #sheetの日にち取得
    date_name = os.path.splitext(os.path.basename(xlfile_path_list[n]))[0]
    year = int(date_name[0:4])
    month = int(date_name[4:6])
    day = int(date_name[6:8])
    #取得した日にちに１日足す
    next_date_name = datetime.date(year, month, day) + datetime.timedelta(days=1)
    next_date_name = next_date_name.strftime("%Y%m%d")
    
    #いらない栽培ベッドデータ削除
    for column in range(1, sheet.max_column):
        value = sheet.cell(row=1, column=column).value
        if value not in (4, 10, 16, 22, 27, None):
            sheet.delete_cols(column, 2)
    
    #初日のデータ        
    if n == 0:
        for row in range(1, sheet.max_row+1):
            for column in range(1, sheet.max_column+1):
                value_init = sheet.cell(row=row, column=column).value
                #初日のシートをsheet_initにコピー
                sheet_init.cell(row=row, column=column).value = value_init
                sheet_init_2.cell(row=row, column=column).value = value_init
                sheet_uekae_write.cell(row=row, column=column).value = value_init
                
                sheet_tsubomi.cell(row=row, column=column).value = value_init
                sheet_flower.cell(row=row, column=column).value = value_init
                sheet_uekae.cell(row=row, column=column).value = value_init
                #出蕾日、開花日の欄の範囲のとき
                if (3 <= row) and (2 <= column):
                    if sheet.cell(row=row, column=column).value != None:
                        sheet_tsubomi.cell(row=row, column=column).value = 0
                        sheet_flower.cell(row=row, column=column).value = 0
                        sheet_uekae.cell(row=row, column=column).value = 0
                        sheet_init_2.cell(row=row, column=column).value = 0
                        sheet_uekae_write.cell(row=row, column=column).value = 0
                    #1のとき
                    elif sheet.cell(row=row, column=column).value == 1:
                        #sheet_initの1を日にちに置き換える
                        sheet_init.cell(row=row, column=column).value = int(date_name)
                        #各シートの開花日を赤
                        if row % 2 == 1:
                            sheet.cell(row=row, column=column).fill = get_green()
                        #各シートの出蕾日を緑
                        else:
                            sheet.cell(row=row, column=column).fill = get_red()
                            
                if sheet.cell(row=row, column=column).has_style:
                    sheet_pre.cell(row=row, column=column).fill = sheet.cell(row=row, column=column).fill._StyleProxy__target
    #２日目以降のデータ                        
    else:
        for row in range(3, sheet.max_row+1):
            for column in range(2, sheet.max_column+1):
                
                value_init = sheet_init.cell(row=row, column=column).value
                value = sheet.cell(row=row, column=column).value
                
                #値がNoneなら前日の値を引き継ぐ
                if value == None:
                    sheet.cell(row=row, column=column).value = sheet_pre.cell(row=row, column=column).value
                if row % 2 == 1:
                    #つぼみデータ行と花データ行の値を取得
                    value_tsubomi = sheet.cell(row=row, column=column).value
                    value_flower = sheet.cell(row=row+1, column=column).value
                    
                    value_pre_tsubomi = sheet_pre.cell(row=row, column=column).value
                    value_pre_flower = sheet_pre.cell(row=row+1, column=column).value
                    
                    #つぼみデータが１のとき
                    if value_tsubomi == 1:
                        #調査票を緑に塗る
                        # green = get_green()
                        sheet.cell(row=row, column=column).fill = get_green()
                        #前日つぼみが出ていなければ、結果表に出蕾日を記録
                        if value_pre_tsubomi == 0:
                            sheet_init.cell(row=row, column=column).value = int(date_name)
                            
                    elif value_tsubomi == 2:
                        sheet.cell(row=row, column=column).fill = get_green(value_tsubomi*70)
                        #前日つぼみが出ていなければ、結果表に出蕾日を記録
                        if value_pre_tsubomi == 0:
                            sheet_init_2.cell(row=row, column=column).value = int(date_name)
                            
                    #花データが１のとき
                    elif value_flower == 1:
                        #調査票を赤に塗る
                        sheet.cell(row=row+1, column=column).fill = get_red()
                        #前日花が咲いていなければ、結果表に開花日を記録
                        if value_pre_flower == 0:
                            sheet_init.cell(row=row+1, column=column).value = int(date_name)
                            
                    elif value_flower == 2:
                        #調査票を赤に塗る
                        sheet.cell(row=row+1, column=column).fill = get_red(value_flower*70)
                        #前日花が咲いていなければ、結果表に開花日を記録
                        if value_pre_flower == 0:
                            sheet_init_2.cell(row=row+1, column=column).value = int(date_name)
                    
                            
                    #両方０のとき前日のつぼみか花が０でないなら調査票を青に塗る（植え替え）、結果表の日にちを０に戻す       
                    if value_tsubomi == value_flower == 0:
                        sheet.cell(row=row, column=column).fill = sheet_pre.cell(row=row, column=column).fill._StyleProxy__target
                        if value_pre_tsubomi != 0 or value_pre_flower != 0:
                            sheet_init.cell(row=row, column=column).value = 0
                            sheet_init.cell(row=row+1, column=column).value = 0
                            sheet.cell(row=row, column=column).fill = get_blue()
                            sheet_uekae_write.cell(row=row, column=column).value = int(date_name)
                            
    for row in range(3, sheet.max_row+1):
        for column in range(2, sheet.max_column+1):
            sheet_pre.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
            sheet_pre.cell(row=row, column=column).fill = sheet.cell(row=row, column=column).fill._StyleProxy__target


#３行目からは
for row in range(3, sheet.max_row+1):
    for column in range(1, sheet.max_column+1):
        value_init = sheet_init.cell(row=row, column=column).value
        value_init_2 = sheet_init_2.cell(row=row, column=column).value
        if column >= 2:
            #調査票の中身を空欄に
            sheet.cell(row=row, column=column).value = None
        #tsubomi用にコピー
        if row % 2 == 1:
            sheet_tsubomi.cell(row=(row//2)+2, column=column).value = value_init
            sheet_tsubomi_2.cell(row=(row//2)+2, column=column).value = value_init_2
            sheet_uekae.cell(row=(row//2)+2, column=column).value = sheet_uekae_write.cell(row=row, column=column).value
        #flower用にコピー
        else:
            sheet_flower.cell(row=row//2+1, column=column).value = value_init
            sheet_flower_2.cell(row=row//2+1, column=column).value = value_init_2                    

#保存_調査票
wb.save(save_tyousa_path + "\\" + str(next_date_name) + ".xlsx")

#保存_結果表（出力日は、エクセルリストの最後の日付）
wb_init.save(save_result_path + "\\" + str(date_name) + "_result.xlsx")
wb_init_2.save(save_result_path + "\\" + str(date_name) + "_result_02.xlsx")
wb_tsubomi.save(save_result_path + "\\" + str(date_name) + "_result_tsubomi.xlsx")
wb_flower.save(save_result_path + "\\" + str(date_name) + "_result_flower.xlsx")
wb_tsubomi_2.save(save_result_path + "\\" + str(date_name) + "_result_tsubomi_02.xlsx")
wb_flower_2.save(save_result_path + "\\" + str(date_name) + "_result_flower_02.xlsx")
wb_uekae.save(save_result_path + "\\" + str(date_name) + "_result_uekae.xlsx")