#モジュールのインポート
import openpyxl as px
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os
import glob
import datetime

#エクセルファイル読み込み
###############result_anarysysのフォルダのパスに書き換えてください##################
xlfile_folder_path = r"D:\D_tokudome\D_desktop\result_analysis\*.xlsx"
xlfile_path_list = glob.glob(xlfile_folder_path)

###############保存先のパスに書き換えてください###################
save_path = r"D:\D_tokudome\D_desktop\result3"
os.makedirs(save_path, exist_ok=True)

#色指定
fill_red = PatternFill(patternType="solid", fgColor="FF0000", bgColor="FF0000")
fill_green = PatternFill(patternType="solid", fgColor="008000", bgColor="008000")
fill_blue = PatternFill(patternType="solid", fgColor="0000FF", bgColor="0000FF")

#フォント指定
font = Font(name="Yu Gothic")

#空のワークブック作成　result書き込み用
wb_init = px.Workbook()
wb_init["Sheet"].title = "Sheet1"
sheet_init = wb_init["Sheet1"]

#出蕾日、開花日記録、色塗り
for n in range(0, len(xlfile_path_list)):
    print(xlfile_path_list[n])
    #現在と１つ前のワークブック作成
    pre_wb = px.load_workbook(xlfile_path_list[n-1])
    pre_sheet = pre_wb["Sheet1"]
    wb = px.load_workbook(xlfile_path_list[n])
    sheet = wb["Sheet1"]
    
    #sheetの日にち取得
    date_name = os.path.splitext(os.path.basename(xlfile_path_list[n]))[0]
    year = int(date_name[0:4])
    month = int(date_name[4:6])
    day = int(date_name[6:8])
    #取得した日にちに１日足す
    next_date_name = datetime.date(year, month, day) + datetime.timedelta(days=1)
    next_date_name = next_date_name.strftime("%Y%m%d")
    
    #いらない栽培ベッドデータ削除
    for column in range(1, sheet.max_column):
        value = sheet.cell(row=1, column=column).value
        pre_value = pre_sheet.cell(row=1, column=column).value
        if value not in (4, 10, 16, 22, 27, None):
            # pre_sheet.cell(row=1, column=column).fill = fill_blue
            # print(pre_sheet.cell(row=1, column=column).fill.bgColor.value)
            sheet.delete_cols(column, 2)
            pre_sheet.delete_cols(column, 2)
            
            
    print(sheet.max_row)
    print(sheet.max_column)
            
    if n == 0:
        for row in range(1, sheet.max_row+1):
            for column in range(1, sheet.max_column+1):
                #初日のシートをsheet_initにコピー
                sheet_init.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
                # sheet_init.cell(row=row, column=column).font = font
                #出蕾日、開花日の欄の範囲のとき
                if (3 <= row <= 300) and (2 <= column <= 21):
                    #空白は0で埋める
                    # if sheet.cell(row=row, column=column).value == None:
                    #     sheet_init.cell(row=row, column=column).value = 0
                    #1のとき
                    if sheet.cell(row=row, column=column).value == 1:
                        #sheet_initの1を日にちに置き換える
                        sheet_init.cell(row=row, column=column).value = date_name
                        #各シートの開花日を赤
                        if row % 2 == 1:
                            sheet.cell(row=row, column=column).fill = fill_green
                        #各シートの出蕾日を緑
                        else:
                            sheet.cell(row=row, column=column).fill = fill_red
                            
    else:
        for row in range(3, sheet.max_row+1):
            for column in range(2, sheet.max_column+1):
                
                value_init = sheet_init.cell(row=row, column=column).value
                value = sheet.cell(row=row, column=column).value
                pre_value = pre_sheet.cell(row=row, column=column).value
                
                #値がNoneなら前日の値を引き継ぐ
                if value == None:
                    sheet.cell(row=row, column=column).value = pre_sheet.cell(row=row, column=column).value
                #値が0なら日付を０に戻す
                if value == 0:
                    sheet_init.cell(row=row, column=column).value = 0
                    
                if row % 2 == 1:
                    #つぼみデータ行と花データ行の値を取得
                    value_init_tsubomi = sheet_init.cell(row=row, column=column).value
                    value_init_flower = sheet_init.cell(row=row+1, column=column).value
                    
                    value_tsubomi = sheet.cell(row=row, column=column).value
                    value_flower = sheet.cell(row=row+1, column=column).value
                    
                    pre_value_tsubomi = pre_sheet.cell(row=row, column=column).value
                    pre_value_flower = pre_sheet.cell(row=row+1, column=column).value
                    
                    
                    #つぼみデータが１のとき
                    if value_tsubomi in (1, 9):
                        sheet.cell(row=row, column=column).fill = fill_green
                        #前日つぼみが出ていなければ、出蕾日を記録
                        if pre_value_tsubomi not in (1, 9):
                            sheet_init.cell(row=row, column=column).value = int(date_name)
                            
                    #花データが１のとき
                    elif value_flower in (1, 9):
                        sheet.cell(row=row+1, column=column).fill = fill_red
                        #前日花が咲いていなければ、開花日を記録
                        if pre_value_flower not in (1, 9):
                            sheet_init.cell(row=row+1, column=column).value = int(date_name)
                    
                            
                    #両方０のとき、もしくは前日が青色であれば青色に塗る（植え替え）       
                    if value_tsubomi == value_flower == 0:
                        if (pre_value_tsubomi != 0 or pre_value_flower != 0) or (pre_sheet.cell(row=row, column=column).fill.bgColor.value == "000000FF"):
                            sheet.cell(row=row, column=column).fill = fill_blue
                    # elif (pre_value_tsubomi in (1, 9) or pre_value_flower in (1, 9)) or (pre_sheet.cell(row=row, column=column).fill == fill_blue):
                    #     if sheet.cell(row=row, column=column).value != None:
                    #         sheet.cell(row=row, column=column).fill = fill_blue
                            
                    
            
    wb.save(save_path + "\\" + str(next_date_name) + ".xlsx")       
    
    
    
    
    
    
    
    
    
    
#     if n == 0:
#         for row in range(1, 301):
#             for column in range(1, 22):
#                 #初日のシートをsheet_initにコピー
#                 sheet_init.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
#                 # sheet_init.cell(row=row, column=column).font = font
#                 #出蕾日、開花日の欄の範囲のとき
#                 if (3 <= row <= 300) and (2 <= column <= 21):
#                     #空白は0で埋める
#                     if sheet.cell(row=row, column=column).value == None:
#                         sheet_init.cell(row=row, column=column).value = 0
#                     #1のとき
#                     elif sheet.cell(row=row, column=column).value == 1:
#                         #sheet_initの1を日にちに置き換える
#                         sheet_init.cell(row=row, column=column).value = date_name
#                         #各シートの開花日を赤
#                         if row % 2 == 1:
#                             sheet.cell(row=row, column=column).fill = fill_green
#                         #各シートの出蕾日を緑
#                         else:
#                             sheet.cell(row=row, column=column).fill = fill_red
                            
#     else:
#         for row in range(3, 301):
#             for column in range(2, 22):
#                 if row % 2 == 1:
#                     #つぼみデータ行と花データ行の値を取得
#                     value_init_tsubomi = sheet_init.cell(row=row, column=column).value
#                     value_init_flower = sheet_init.cell(row=row+1, column=column).value
                    
#                     value_tsubomi = sheet.cell(row=row, column=column).value
#                     value_flower = sheet.cell(row=row+1, column=column).value
                    
#                     pre_value_tsubomi = pre_sheet.cell(row=row, column=column).value
#                     pre_value_flower = pre_sheet.cell(row=row+1, column=column).value

#                     #つぼみデータが１のとき
#                     if value_tsubomi == 1:
#                         sheet.cell(row=row, column=column).fill = fill_green
#                         #前日つぼみが出ていなければ、出蕾日を記録
#                         if pre_value_tsubomi != 1:
#                             sheet_init.cell(row=row, column=column).value = int(date_name)
                            
#                     #花データが１のとき
#                     elif value_flower == 1:
#                         sheet.cell(row=row+1, column=column).fill = fill_red
#                         #前日花が咲いていなければ、開花日を記録
#                         if pre_value_flower != 1:
#                             sheet_init.cell(row=row+1, column=column).value = int(date_name)
                            
#                     #両方０のとき、もしくは前日が青色であれば青色に塗る（植え替え）       
#                     elif (pre_value_tsubomi == 1 or pre_value_flower == 1) or (pre_sheet.cell(row=row, column=column).fill == fill_blue):
#                         if sheet.cell(row=row, column=column).value != None:
#                             sheet.cell(row=row, column=column).fill = fill_blue
                            
#                     if value_tsubomi == 0:
#                         sheet_init.cell(row=row, column=column).value = 0
#                     if value_flower == 0:
#                         sheet_init.cell(row=row+1, column=column).value = 0
    
#     #保存_調査票（出力日は参照したエクセルの日付の次の日）                        
#     wb.save(save_path + "\\" + str(next_date_name) + ".xlsx")
                        
# #出蕾日と開花日でブックを分割                    
# wb_tsubomi = px.Workbook()
# sheet_tsubomi = wb_tsubomi["Sheet"]

# wb_flower = px.Workbook()
# sheet_flower = wb_flower["Sheet"]

# for row in range(1, 3):
#     for column in range(1, 22):
#         sheet_tsubomi.cell(row=row, column=column).value = sheet_init.cell(row=row, column=column).value
#         sheet_flower.cell(row=row, column=column).value = sheet_init.cell(row=row, column=column).value
# for row in range(3, 301):
#     for column in range(1, 22):
#         if row % 2 == 0:
#             sheet_flower.cell(row=(row/2)+2, column=column).value = sheet_init.cell(row=row, column=column).value
#         else:
#             sheet_tsubomi.cell(row=(row//2)+2, column=column).value = sheet_init.cell(row=row, column=column).value                    

# #保存_結果表（出力日は、エクセルリストの最後の日付）
# wb_init.save(save_path + "\\" + str(date_name) + "_result.xlsx")
# wb_tsubomi.save(save_path + "\\" + str(date_name) + "_result_tsubomi.xlsx")
# wb_flower.save(save_path + "\\" + str(date_name) + "_result_flower.xlsx")