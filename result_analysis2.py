#モジュールのインポート
import openpyxl as px
import os
import glob

#エクセルファイル読み込み
xlfile_folder_path = r"D:\D_tokudome\D_desktop\result_analysis\*.xlsx"
xlfile_path_list = glob.glob(xlfile_folder_path)
os.makedirs(r"D:\D_tokudome\D_desktop\result", exist_ok=True)

for n, f in enumerate(xlfile_path_list):
    xlfile_path = f
    wb = px.load_workbook(xlfile_path)
    date_name = os.path.splitext(os.path.basename(xlfile_path))[0]
    if n == 0:
        wb_init = wb
        print(date_name)
    print(f)
    sheet = wb["Sheet1"]
    sheet_init = wb_init["Sheet1"]
    for row in range(3, 301, 2):
        for column in range(2, 22):
            value = sheet.cell(row=row, column=column).value
            if value == 1:
                sheet_init.cell(row=row, column=column).value = int(date_name)-1
                sheet_init.cell(row=row, column=column).value += 1

wb_init.save(r"D:\D_tokudome\D_desktop\result\\" + str(date_name) + ".xlsx")

    # #ファイル名から日にち取得
    # date_name = os.path.splitext(os.path.basename(xlfile_path))[0]
    # print(date_name)

    # #エクセルファイルコピー
    # wb_copy = wb

    # #出蕾日行列取得
    # sheet = wb["Sheet1"]
    # sheet_copy = wb_copy["Sheet1"]
    # for row in range(3, 301, 2):
    #     for column in range(2, 22):
    #         value = sheet.cell(row=row, column=column).value
    #         if value == 1:
    #             sheet_copy.cell(row=row, column=column).value = int(date_name)
    #             print("row:", row, "column:", column, "\n")



#１を出蕾日に書き換え


#保存
# wb_copy.save("result.xlsx")
